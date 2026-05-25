import os
from decimal import Decimal
import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from billing.models import Item, Shop, UserProfile
from billing.sarvam_helper import generate_unique_barcode

class Command(BaseCommand):
    help = "Seed master inventory from the local Excel file and create the admin user account"

    def handle(self, *args, **options):
        # 1. Seed the Admin User
        admin_email = "vyapaarbills@gmail.com"
        admin_pass = "Sainit@9192"
        self.stdout.write(f"Checking for admin account: {admin_email}...")
        
        user, created = User.objects.get_or_create(
            username=admin_email,
            defaults={
                "email": admin_email,
                "is_superuser": True,
                "is_staff": True
            }
        )
        user.set_password(admin_pass)
        user.is_superuser = True
        user.is_staff = True
        user.save()
        
        # Ensure profile exists
        UserProfile.objects.update_or_create(
            user=user,
            defaults={
                "phone_number": "8888888888",
                "annual_income": Decimal("50000000.00"),
                "plan": UserProfile.BASIC,
                "custom_plan_price": None,
                "custom_monthly_runs": None,
                "is_active": True
            }
        )
        if created:
            self.stdout.write(self.style.SUCCESS(f"Created admin account {admin_email}"))
        else:
            self.stdout.write(self.style.SUCCESS(f"Updated existing admin account {admin_email}"))

        # 2. Seed Master Inventory from Excel
        excel_path = r"C:\Users\sahit\OneDrive\Desktop\software\General_Store_Inventory.xlsx"
        fallback_path = os.path.join(settings.BASE_DIR.parent, "Supermarket_Unique_Master_Data_Cleaned.xlsx")
        if not os.path.exists(excel_path) and os.path.exists(fallback_path):
            excel_path = fallback_path
        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR(f"Excel file not found at: {excel_path}"))
            return

        self.stdout.write(f"Loading master inventory Excel file from: {excel_path}...")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active

        # Get existing barcodes from DB to ensure uniqueness
        existing_barcodes = set(Item.objects.filter(barcode__isnull=False).values_list("barcode", flat=True))

        created_count = 0
        updated_count = 0

        # Excel Headers: ['S.No', 'Item Name', 'Category', 'HSN/SAC Code', 'Unit of Measure', 'Selling Price (Rs)', 'MRP (Rs)', 'Margin (%)', 'GST %', 'Stock Qty', 'Net Worth (Rs)', 'BARCODE']
        # Row 1 is header, data starts at Row 2
        for row_idx in range(2, sheet.max_row + 1):
            item_name = sheet.cell(row=row_idx, column=2).value
            category = sheet.cell(row=row_idx, column=3).value
            hsn_code = sheet.cell(row=row_idx, column=4).value
            uom = sheet.cell(row=row_idx, column=5).value
            gst_val = sheet.cell(row=row_idx, column=9).value
            barcode_val = sheet.cell(row=row_idx, column=12).value

            if not item_name:
                continue

            # Beautify fields
            full_item_name = str(item_name).strip()
            item_cat = str(category).strip() if category else "Grocery"
            item_hsn = str(hsn_code).strip() if hsn_code else "9999"
            item_uom = str(uom).strip() if uom else "pcs"

            # Parse GST
            gst_percent = Decimal("18.00")
            if gst_val is not None:
                try:
                    gst_str = str(gst_val).replace("%", "").strip()
                    gst_percent = Decimal(gst_str)
                except Exception:
                    pass

            # Check if global item already exists
            item = Item.objects.filter(name=full_item_name, is_global=True).first()

            if item:
                # Update attributes
                item.category = item_cat
                item.hsn_code = item_hsn
                item.unit_of_measure = item_uom
                item.gst_percent = gst_percent
                item.save()
                updated_count += 1
            else:
                # Parse or generate barcode
                barcode = str(barcode_val).strip() if barcode_val else ""
                if not barcode or barcode.lower() in ("none", "null", ""):
                    barcode = generate_unique_barcode()
                
                # Check uniqueness again
                if barcode in existing_barcodes:
                    barcode = generate_unique_barcode()
                existing_barcodes.add(barcode)

                Item.objects.create(
                    name=full_item_name,
                    category=item_cat,
                    hsn_code=item_hsn,
                    unit_of_measure=item_uom,
                    barcode=barcode,
                    mrp=Decimal("100.00"),  # Standard default MRP for master items
                    gst_percent=gst_percent,
                    is_global=True,
                    shop_type=Shop.GENERAL,
                    is_active=True
                )
                created_count += 1

            if (row_idx - 1) % 200 == 0:
                self.stdout.write(f"Processed {row_idx - 1} rows...")

        self.stdout.write(self.style.SUCCESS(
            f"Excel Seeding Complete! Created {created_count} new master items, updated {updated_count} items."
        ))
