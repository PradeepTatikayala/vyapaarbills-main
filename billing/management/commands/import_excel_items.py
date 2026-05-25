import os
import random
from decimal import Decimal
import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from billing.models import Item, Shop

def calculate_ean13_checksum(code_str):
    sum_even = sum(int(code_str[i]) for i in range(1, 12, 2))
    sum_odd = sum(int(code_str[i]) for i in range(0, 12, 2))
    total = sum_odd + sum_even * 3
    checksum = (10 - (total % 10)) % 10
    return str(checksum)

def generate_barcode(existing_barcodes):
    while True:
        # India retail barcode prefix is 890
        base = "890" + "".join(str(random.randint(0, 9)) for _ in range(9))
        checksum = calculate_ean13_checksum(base)
        barcode = base + checksum
        if barcode not in existing_barcodes:
            existing_barcodes.add(barcode)
            return barcode

class Command(BaseCommand):
    help = "Import master items from Excel spreadsheet"

    def handle(self, *args, **options):
        excel_path = os.path.join(settings.BASE_DIR, "..", "Supermarket_Unique_Master_Data_Cleaned.xlsx")
        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR(f"Excel file not found at: {excel_path}"))
            return

        self.stdout.write(f"Loading Excel file from: {excel_path}...")
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        # Get existing barcodes from DB to ensure uniqueness
        existing_barcodes = set(Item.objects.filter(barcode__isnull=False).values_list("barcode", flat=True))
        
        self.stdout.write("Seeding master items...")
        created_count = 0
        updated_count = 0

        # Excel Columns: Category, Brand, Item Name, Variant, HSN Code, GST %, MRP ()
        for row_idx in range(2, sheet.max_row + 1):
            category = sheet.cell(row=row_idx, column=1).value
            brand = sheet.cell(row=row_idx, column=2).value
            raw_name = sheet.cell(row=row_idx, column=3).value
            variant = sheet.cell(row=row_idx, column=4).value
            hsn_code = sheet.cell(row=row_idx, column=5).value
            gst_val = sheet.cell(row=row_idx, column=6).value
            mrp_val = sheet.cell(row=row_idx, column=7).value

            if not raw_name:
                continue

            # Formulate the item name beautifully
            name_parts = []
            if brand and str(brand).strip().lower() != 'none':
                name_parts.append(str(brand).strip())
            name_parts.append(str(raw_name).strip())
            if variant and str(variant).strip().lower() != 'none':
                name_parts.append(str(variant).strip())
            
            full_item_name = " ".join(name_parts)

            # Clean category
            item_cat = str(category).strip() if category else "Grocery"
            
            # Clean HSN code
            item_hsn = str(hsn_code).strip() if hsn_code else "1006"

            # Parse GST %
            if gst_val:
                gst_str = str(gst_val).replace("%", "").strip()
                try:
                    gst_percent = Decimal(gst_str)
                except Exception:
                    gst_percent = Decimal("0.00")
            else:
                gst_percent = Decimal("5.00")

            # Parse MRP
            mrp = None
            if mrp_val and str(mrp_val).strip().lower() != 'none':
                try:
                    mrp = Decimal(str(mrp_val).strip())
                except Exception:
                    pass

            # Try to get existing item by name or barcode
            # Check if there is an item with the same name and category
            item = Item.objects.filter(name=full_item_name, category=item_cat, is_global=True).first()
            
            if item:
                # Update item fields
                item.hsn_code = item_hsn
                item.gst_percent = gst_percent
                if mrp is not None:
                    item.mrp = mrp
                item.save()
                updated_count += 1
            else:
                # Generate unique barcode
                barcode = generate_barcode(existing_barcodes)
                
                Item.objects.create(
                    name=full_item_name,
                    category=item_cat,
                    hsn_code=item_hsn,
                    unit_of_measure="pcs" if "kg" not in full_item_name.lower() and "gm" not in full_item_name.lower() and "ml" not in full_item_name.lower() and "l " not in full_item_name.lower() else ("kg" if "kg" in full_item_name.lower() else "pcs"),
                    barcode=barcode,
                    mrp=mrp or Decimal("100.00"),  # fallback MRP
                    gst_percent=gst_percent,
                    is_global=True,
                    shop_type=Shop.GENERAL,
                    is_active=True
                )
                created_count += 1

        # Seed additional items for ReadyMade and Fancy categories
        self.seed_readymade_and_fancy_items(existing_barcodes)

        self.stdout.write(self.style.SUCCESS(f"Import complete! Created {created_count} new master items, updated {updated_count} items."))

    def seed_readymade_and_fancy_items(self, existing_barcodes):
        # ReadyMade items
        readymade_items = [
            ("Mens Cotton Shirt", "Readymade / Mens Wear", "6205", 5, 899.00),
            ("Mens Formal Trouser", "Readymade / Mens Wear", "6203", 12, 1299.00),
            ("Womens Kurti", "Readymade / Womens Wear", "6206", 5, 999.00),
            ("Kids T-Shirt", "Readymade / Kids Wear", "6111", 5, 349.00),
            ("Premium Silk Saree", "Readymade / Womens Wear", "5007", 12, 3499.00),
            ("Leather Belt", "Readymade / Accessories", "4203", 5, 499.00),
            ("Premium Handbag", "Readymade / Accessories", "4202", 12, 2999.00),
        ]

        for name, cat, hsn, gst, mrp in readymade_items:
            if not Item.objects.filter(name=name, is_global=True).exists():
                barcode = generate_barcode(existing_barcodes)
                Item.objects.create(
                    name=name,
                    category=cat,
                    hsn_code=hsn,
                    unit_of_measure="pcs",
                    barcode=barcode,
                    mrp=Decimal(mrp),
                    gst_percent=Decimal(gst),
                    is_global=True,
                    shop_type=Shop.READYMADE,
                    is_active=True
                )

        # Fancy items
        fancy_items = [
            ("Matte Lipstick", "Fancy / Cosmetics", "3304", 18, 299.00),
            ("Nail Polish Combo", "Fancy / Cosmetics", "3304", 18, 199.00),
            ("Designer Necklace Set", "Fancy / Jewellery", "7117", 3, 599.00),
            ("Glass Bangles Box", "Fancy / Jewellery", "7018", 3, 150.00),
            ("Teddy Bear 3 Feet", "Fancy / Toys", "9503", 12, 799.00),
            ("Premium Perfume Spray", "Fancy / Cosmetics", "3303", 18, 499.00),
        ]

        for name, cat, hsn, gst, mrp in fancy_items:
            if not Item.objects.filter(name=name, is_global=True).exists():
                barcode = generate_barcode(existing_barcodes)
                Item.objects.create(
                    name=name,
                    category=cat,
                    hsn_code=hsn,
                    unit_of_measure="pcs",
                    barcode=barcode,
                    mrp=Decimal(mrp),
                    gst_percent=Decimal(gst),
                    is_global=True,
                    shop_type=Shop.FANCY,
                    is_active=True
                )
